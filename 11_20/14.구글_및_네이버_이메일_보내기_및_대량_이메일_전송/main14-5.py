"""
엑셀 파일에서 대량으로 이메일 보내는 코드 만들기
"""

from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import getpass

load_wb = load_workbook("11_20/14.구글_및_네이버_이메일_보내기_및_대량_이메일_전송/이메일주소.xlsx", data_only=True)
load_ws = load_wb.active

for i in range(1, load_ws.max_row + 1):
    recv_email_value = load_ws.cell(i, 1).value
    send_email = "123wodnr@naver.com"
    send_pwd = getpass.getpass('Password:')

    try:

        recv_email = recv_email_value

        smtp_name = "smtp.naver.com"
        smtp_port = 587

        msg = MIMEMultipart()

        msg['Subject'] = f"엑셀에서 수신주소를 읽어 보내는 테스트: {send_email} > {recv_email}"
        msg['From'] = send_email
        msg['To'] = recv_email

        text = """
                메일내용 입니다.
                감사합니다.
                """

        msg.attach(MIMEText(text))

        s = smtplib.SMTP(smtp_name, smtp_port)
        s.starttls()
        s.login(send_email, send_pwd)
        s.sendmail(send_email, recv_email, msg.as_string())
        print("성공:", recv_email_value)
        s.quit()
    except:
        print("에러:", recv_email_value)
