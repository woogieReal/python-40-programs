"""
오픈 API를 이용해 주소를 좌표로 변환하여 엑셀파일로 생성하는 코드 만들기
"""
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re
from dotenv import load_dotenv
import os

load_dotenv()

file_path = '21_30/27.전국의_대학교_위치_시각화하기/고등교육기관_하반기_주소록-2022.xlsx'
df_from_excel = pd.read_excel(file_path, engine='openpyxl')

# 5번째 위치의 데이터를 columns으로 설정
df_from_excel.columns = df_from_excel.loc[4].tolist()
# -> ['연도', '학교종류', '시도', '행정구', '학교명', '학교명(영문)', '본분교', '학교상태', '설립', '우편번호', '주소', '전화번호', '팩스번호', '홈페이지']

# 0~5줄의 데이터를 버린다. 설명이나 이름등의 데이터 제거
df_from_excel = df_from_excel.drop(index=list(range(0, 5)))


# API 접속내용
url = 'http://api.vworld.kr/req/address?'
params = 'service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=false&format=json&type='
road_type = 'ROAD'  # 도로명주소
road_type2 = 'PARCEL'  # 지번주소
address = '&address='
keys = '&key='

# 발급받은 인증키
primary_key = os.environ.get('GEO_DEV_KEY')


# 주소를 x, y 좌표로 반환해주는 함수
def request_geo(road):
    page = requests.get(url+params+road_type+address+road+keys+primary_key)
    json_data = page.json()
    if json_data['response']['status'] == 'OK':
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        return x, y
    else:
        x = 0
        y = 0
        return x, y


# 엑셀 파일이 읽어오고 없으면 만든다.
try:
    wb = load_workbook('21_30/27.전국의_대학교_위치_시각화하기/학교주소좌표.xlsx', data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active

university_list = df_from_excel['학교명'].to_list()
address_list = df_from_excel['주소'].to_list()

for num, value in enumerate(address_list):
    # 주소에서 () 괄효 부분을 제거
    addr = re.sub(r'\([^)]*\)', '', value)
    print(addr)

    # API를 활용하여 주소를 좌표로 변환
    x, y = request_geo(addr)

    # 학교명, 주소, x, y의 순서대로 엑셀에 저장
    sheet.append([university_list[num], addr, x, y])


wb.save('21_30/27.전국의_대학교_위치_시각화하기/학교주소좌표.xlsx')
