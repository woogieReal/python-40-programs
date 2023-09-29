"""
판다스 라이브러리로 값을 엑셀로 저장 후 불러오는 코드 만들기
"""

from openpyxl import load_workbook # 엑셀에서 파일을 읽이 위해서 openpyxl에서 load_workbook을 불러옴
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# 엑셀 파일을 읽어옴
load_xb = load_workbook("수료증명단.xlsx")

# 읽어온 엑셀 파일에서 활성화된 시트를 불러온다.
load_ws = load_xb.active

name_list = []
birthday_list = []
ho_list = []

for i in range(1, load_ws.max_row + 1):
    name_list.append(load_ws.cell(i, 1).value)
    birthday_list.append(load_ws.cell(i, 2).value)
    ho_list.append(load_ws.cell(i, 3).value)

print(name_list)
print(birthday_list)
print(ho_list)
