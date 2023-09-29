"""
수집한 이메일 주소를 엑셀에 저장하는 코드 만들기
"""

import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

url = 'https://news.v.daum.net/v/20211129144552297'

headers = {
        'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'text/html; charset=utf-8'
        }

response = requests.get(url, headers=headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

results = list(set(results))

print(results)

try:
    wb = load_workbook("email.xlsx", data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active

for result in results:
    sheet.append([result])

wb.save("email.xlsx")
